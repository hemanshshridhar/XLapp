from langchain_community.llms import Ollama
import streamlit as st
import os
from dotenv import load_dotenv
import json
import re
import tempfile
from openpyxl import load_workbook
from utils import  converter


os.environ["LANGCHAIN_TRACING_V2"] = "true"
api_key = st.secrets["OPENAI_API_KEY"]
langchain_key = st.secrets["LANGCHAIN_API_KEY"]

def clean_and_parse_json(response):
    """Cleans and parses JSON safely, handling errors."""
    try:
        if not response or not isinstance(response, str):
            raise ValueError("Response is empty or not a string.")


        response = re.sub(r"//.*", "", response)

        response = response.replace("'", '"')  

   
        match = re.search(r"\{.*\}", response, re.DOTALL)
        if match:
            return json.loads(match.group(0)) 

    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
    except ValueError as e:
        print(f"Value error: {e}")

    return None 

def update_excel_mapping(data_dict, user_query):
    prompt = ChatPromptTemplate.from_template("""
    You are analyzing structured data extracted from an Excel sheet.
    The data is provided in dictionary format, where keys represent the content in a cell 
    and values indicate corresponding cell address.
    Additionally, a **user query** is provided. Your task is to:
    1. Identify and count the **distinct tables** in the sheet based on relative address differences.
    2. Find patterns within the data.

    #### Given Data:
    - **Excel Data (Dictionary Format):**  
    {data_dict}

    - **User Query:**  
    {user_query}

    ### **Expected Output Format:**  
    ```plaintext
    Number of tables: [total_count]
    Give a textual description for the input data 
    ```
    """)

    response = model.invoke(prompt.format(
    data_dict=json.dumps(data_dict, indent=4),
    user_query=user_query
    ))

    response_text = response.content
    st.write("Response:", response_text)
    return response



def update_excel_from_json(json_data, excel_file, output_file, sheet_name):
  
    wb = load_workbook(excel_file, keep_vba=True)


    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        return

    sheet = wb[sheet_name]

  
    for key, cell_refs in json_data.items():

        if isinstance(cell_refs, str):
    
            cell_list = [cell.strip() for cell in cell_refs.split(',')]

            for cell_ref in cell_list:
                try:
                    sheet[cell_ref] = key  
                except ValueError:
                    print(f"Invalid cell reference: {cell_ref}")

        else:

            if isinstance(cell_refs, str):
                cell_list = [cell.strip() for cell in cell_refs.split(',')]
                for cell_ref in cell_list:
                    try:
                        sheet[cell_ref] = key  
                    except ValueError:
                        print(f"Invalid cell reference: {cell_ref}")


    wb.save(output_file)

st.title("Excel sheet analyser ")
st.write("Upload a `.xlsm` file ")

uploaded_file = st.file_uploader("Upload your .xlsm file", type=["xlsm"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name

    user_command = st.text_input("Enter your query")

    if st.button("Get answer"):
        st.write("Excel sheet analysis")
        if user_command:
            wb = load_workbook(file_path, keep_vba=True) 
            sheet = wb['Model Inputs']
            json_file = converter(sheet)
            sheet_name = "Model Inputs"

            data_dict_fixed = {str(k): v for k, v in json_file.items()}
            answer = update_excel_mapping(json_file, user_command)

            st.write("Excel sheet analysis")
        else :
            st.warning("Please give a query")

