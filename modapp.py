import streamlit as st
import os
import json
import re
import tempfile
import openpyxl
from dotenv import load_dotenv
from langchain.chat_models import ChatOpenAI
from langchain.prompts import ChatPromptTemplate
from openpyxl import load_workbook
from utils import converter
import openai
from langchain.vectorstores import Chroma
from langchain.embeddings import OpenAIEmbeddings


load_dotenv()
os.environ["LANGCHAIN_TRACING_V2"] = "true"
os.environ["LANGCHAIN_API_KEY"] = os.getenv("LANGCHAIN_API_KEY")
openai.api_key =     os.getenv("OPENAI_API_KEY")
model = ChatOpenAI(model_name="gpt-4")
client = openai.Client(api_key=openai.api_key)
embedding_model = OpenAIEmbeddings()
# Function to get sheet names from an Excel file
def get_sheet_names(file_path):
    wb = load_workbook(file_path, keep_vba=True)
    return wb.sheetnames

# Function to parse JSON response
def parse_json_response(response_text):
    cleaned_response = re.sub(r"```(?:json)?\n?|```", "", response_text).strip()
    return json.loads(cleaned_response)

# Function to enforce numeric keys in JSON response
def enforce_numeric_keys(data):
    return {float(k) if k.replace('.', '', 1).isdigit() else k: v for k, v in data.items()}


def store_in_chromadb(country_dict):
    vectordb = Chroma(persist_directory="./chroma_db", embedding_function=embedding_model)

    for country, data in country_dict.items():
        doc_text = convert_dict_to_text(country, data)
        vectordb.add_texts([doc_text], metadatas=[{"label": country}])
    
    vectordb.persist()
# Function to update Excel mapping

def update_excel_mapping(data_dict_fixed):
    # Retrieve stored country_dict from RAG
    retrieved_text = query_rag()

    # Convert retrieved text to a dictionary
    country_dict = {line.split(": ")[0]: line.split(": ")[1] for line in retrieved_text.split("\n")}

    # LLM Prompt (No user query, only dict merging)
    prompt = f"""
    You are given two dictionaries:  
    - `data_dict_fixed` contains **Excel data mappings**.  
    - `country_dict` contains **country-specific values**.  

    **Task**:  
    - Update `data_dict_fixed` using values from `country_dict`.  
    - Only modify values that exist in **both** dictionaries.  
    - Keep all other values in `data_dict_fixed` unchanged.  

    #### Given Dictionaries:
    ```json
    {{
        "data_dict_fixed": {json.dumps(data_dict_fixed, separators=(',', ':'))},
        "country_dict": {json.dumps(country_dict, separators=(',', ':'))}
    }}
    ```

    #### Expected Output:
    Return the updated `data_dict_fixed` **in valid JSON format only**.
    """

    response = client.chat.completions.create(
        model="gpt-4-turbo",
        messages=[{"role": "system", "content": "You are an AI assistant."},
                  {"role": "user", "content": prompt}],
        temperature=0.2
    )

    response_text = response.choices[0].message.content.strip()
    updated_dict = json.loads(response_text)  # Parse LLM's response

    return updated_dict

def query_rag():
    vectordb = Chroma(persist_directory="./chroma_db", embedding_function=embedding_model)
    results = vectordb.similarity_search("country_dict", k=1)  
    return results[0].page_content if results else "{}"
# Example query
# user_query = "Country details"
# retrieved_text = query_rag(user_query)

# print("\n🔍 Retrieved Data:\n", retrieved_text)
# def update_excel_mapping(data_dict, user_query):
#     prompt = f"""
#     You are analyzing structured data extracted from an Excel sheet.
#     Modify the data **precisely as requested** by the user.
#     Maintain the original format, **do not enclose integers in quotes**, and **return only valid JSON**.

#     #### Given Data:
#     ```json
#     {json.dumps(data_dict, separators=(',', ':'))}
#     ```
    
#     #### User Query:
#     {user_query}
    
#     #### Expected Output:
#     Return the updated dictionary **in valid JSON format only**.
#     """

#     response = client.chat.completions.create(
#         model="gpt-4-turbo",
#         messages=[{"role": "system", "content": "You are an AI assistant."},
#                   {"role": "user", "content": prompt}],
#         temperature=0.2
#     )

#     response_text = response.choices[0].message.content.strip()
#     parsed_dict = parse_json_response(response_text)
#     cleaned_dict = enforce_numeric_keys(parsed_dict)

#     st.write("Updated Dictionary:", cleaned_dict)
#     return cleaned_dict

# Function to update Excel sheet based on JSON data
def update_excel_from_json(json_data, excel_file, output_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        st.error(f"Sheet '{sheet_name}' does not exist.")
        return

    sheet = wb[sheet_name]

    for key, cell_refs in json_data.items():
        cell_list = [cell.strip() for cell in cell_refs.split(',')] if isinstance(cell_refs, str) else cell_refs
        value_to_write = key if isinstance(key, (str, bool)) else float(key)

        for cell_ref in cell_list:
            try:
                sheet[cell_ref] = value_to_write
            except ValueError:
                print(f"Invalid cell reference: {cell_ref}")

    wb.save(output_file)
    return f"Updated Excel saved as: {output_file}"

# Streamlit UI
st.title("Excel Modifier App")
st.write("Upload a `.xlsm` file and modify it using natural language.")

uploaded_file = st.file_uploader("Upload your .xlsm file", type=["xlsm"])
uploaded_sheet= st.file_uploader("Upload country sheet", type = ["xlsm"])
if uploaded_file and uploaded_sheet:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as country_temp_file:
        country_temp_file.write(uploaded_country_sheet.read())
        country_file_path = country_temp_file.name        
    # Display available sheet names
    sheet_names = get_sheet_names(file_path)
    selected_sheet = st.selectbox("Select a sheet to modify", sheet_names)

    user_command = st.text_input("Enter your modification command")

    if st.button("Modify Excel"):
        if user_command:
            wb = load_workbook(file_path, keep_vba=True)
            sheet = wb[selected_sheet]
            json_file = converter(sheet)
            data_dict_fixed = {str(k): v for k, v in json_file.items()}
            wb_country = load_workbook(country_file_path, keep_vba=True)
            country_sheet = wb_country[selected_country_sheet]
            country_dict = converter(country_sheet) 
            ## converting the dictioanry to text  that will be embedded in the RAG
            document_text = "\n".join([f"{k}: {v}" for k, v in country_dict.items()])
            # country_data_dict_fixed = {str(k): v for k, v in json_country_file.items()}
            store_in_chromadb(document_text)
            json_updated = update_excel_mapping(data_dict_fixed, user_command)
            output_file_path = file_path.replace(".xlsm", "_modified.xlsm") 
            result = update_excel_from_json(json_updated, file_path, output_file_path, selected_sheet)
            
            st.success(result)
            with open(output_file_path, "rb") as f:
                st.download_button("Download Modified Excel", f, file_name="modified.xlsm", mime="application/vnd.ms-excel")
        else:
            st.warning("Please enter a command to modify the file.")
