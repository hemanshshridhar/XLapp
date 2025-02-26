from langchain_community.llms import Ollama
import streamlit as st
import os
from dotenv import load_dotenv
import json
import re
import tempfile
from openpyxl import load_workbook
from utils import  converter

load_dotenv()
os.environ["LANGCHAIN_TRACING_V2"] = "true"
os.environ["LANGCHAIN_API_KEY"] = os.getenv("LANGCHAIN_API_KEY")


llm = Ollama(model="mistral")
def clean_and_parse_json(response):
    """Cleans and parses JSON safely, handling errors."""
    try:
        if not response or not isinstance(response, str):
            raise ValueError("Response is empty or not a string.")


        response = re.sub(r"//.*", "", response)

        response = response.replace("'", '"')  

   
        match = re.search(r"\{.*\}", response, re.DOTALL)
        if match:
            return json.loads(match.group(0))  # Convert to Python dict

    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
    except ValueError as e:
        print(f"Value error: {e}")

    return None 

def update_excel_mapping(data_dict, user_query):
    prompt = f"""
    You are analyzing structured data extracted from an Excel sheet. The data is provided in dictionary format, where keys represent category names and values indicate corresponding cell locations.

    Additionally, a **user query** is provided. Your task is to:
    1. Identify and count the **distinct tables** in the sheet based on realtive address difference whereever a large differencce in Row number is there, a new table starts .
    2. Ensure that **each table is defined by a set of related rows**.
    3. Find patterns within the data.

    #### Given Data:
    - **Excel Data (Dictionary Format):**  
    {json.dumps(data_dict, indent=4)}
    {print(data_dict)}
    - **User Query:**  
    "{user_query}"

    ### **Expected Output Format:**  
    ```plaintext
    Number of tables: [total_count]
    Table Names: [List of identified table names]

    Provide a concise and accurate response based on the given data.
    """
    response = llm.invoke(prompt) 
    print(response)
    st.write("Raw Response from Ollama:", response)
    # response = clean_and_parse_json(response)
    # response = json.loads(response)
    print(response)
    # # Extract the text response
    # response_text = response["message"]["content"].strip()

    # Extract only the JSON part
    # match = re.search(r"\{.*?\}", response, re.DOTALL)
    # if match:
    #     return json.loads(match.group(0))  # Return parsed JSON
    # return {} 
    return response


def update_excel_from_json(json_data, excel_file, output_file, sheet_name):
    # Load the Excel file and keep the VBA macros if any
    wb = load_workbook(excel_file, keep_vba=True)

    # Access the specific sheet you want to modify
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        return

    sheet = wb[sheet_name]

    # Step through each entry in the JSON data and update corresponding cells
    for key, cell_refs in json_data.items():
        # If the value is a string (cell reference like 'D12' or 'G12, G19, G25')
        if isinstance(cell_refs, str):
            # Split the cell references in case multiple references are provided
            cell_list = [cell.strip() for cell in cell_refs.split(',')]

            for cell_ref in cell_list:
                try:
                    sheet[cell_ref] = key  # Write the key's value to the cell
                except ValueError:
                    print(f"Invalid cell reference: {cell_ref}")

        else:
            # If the value is a number, retrieve corresponding cell references
            if isinstance(cell_refs, str):
                cell_list = [cell.strip() for cell in cell_refs.split(',')]
                for cell_ref in cell_list:
                    try:
                        sheet[cell_ref] = key  # Write the key's value to the cell
                    except ValueError:
                        print(f"Invalid cell reference: {cell_ref}")


    wb.save(output_file)
# Streamlit UI
st.title("Excel sheet analyser ")
st.write("Upload a `.xlsm` file and modify it using natural language.")

uploaded_file = st.file_uploader("Upload your .xlsm file", type=["xlsm"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name

    user_command = st.text_input("Enter your query")

    if st.button("Get answer"):
        st.write("Excel sheet analysis")
        if user_command:
            wb = load_workbook(file_path, keep_vba=True) 
            sheet = wb['Model Inputs']
            json_file = converter(sheet)
            sheet_name = "Model Inputs"
            # print(json_file)
            data_dict_fixed = {str(k): v for k, v in json_file.items()}
            answer = update_excel_mapping(json_file, user_command)
            # output_file_path = file_path.replace(".xlsm", "_modified.xlsm") 
            # result = update_excel_from_json(json_file, file_path,output_file_path , sheet_name)
            st.write("Excel sheet analysis")
        else :
            st.warning("Please give a query")
