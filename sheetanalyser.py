import streamlit as st
import os
from dotenv import load_dotenv
import json
import re
import tempfile
from openpyxl import load_workbook
from utils import  converter,json_to_text
from fastapi import FastAPI
from langchain.prompts import ChatPromptTemplate
from langchain.chat_models import ChatOpenAI
import openai
import openpyxl
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings
from langchain.vectorstores import FAISS
from langchain.text_splitter import CharacterTextSplitter
from langchain.llms import OpenAI
from langchain.chains import RetrievalQA
from langchain.docstore.document import Document
import tiktoken
load_dotenv()
os.environ["LANGCHAIN_TRACING_V2"] = "true"
os.environ["LANGCHAIN_API_KEY"] = os.getenv("LANGCHAIN_API_KEY")
openai.api_key =     os.getenv("OPENAI_API_KEY")
model = ChatOpenAI(model_name="gpt-4")


def clean_and_parse_json(response):
    """Cleans and parses JSON safely, handling errors."""
    try:
        if not response or not isinstance(response, str):
            raise ValueError("Response is empty or not a string.")

 
        response = re.sub(r"//.*", "", response)


        response = response.replace("'", '"')  


        match = re.search(r"\{.*\}", response, re.DOTALL)
        if match:
            return json.loads(match.group(0))  

    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
    except ValueError as e:
        print(f"Value error: {e}")

    return None 



client = openai.Client(api_key=openai.api_key)



def parse_json_response(response_text):
    if not response_text.strip():
        raise ValueError("Error: GPT response is empty. Cannot parse JSON.")


    cleaned_response = re.sub(r"```(?:json)?\n?|```", "", response_text).strip()

    try:
        return json.loads(cleaned_response, parse_float=float, parse_int=int)
    except json.JSONDecodeError as e:
        raise ValueError(f"Error parsing JSON: {e}. Cleaned Response was: {repr(cleaned_response)}")



def enforce_numeric_keys(data):
    """
    Ensures that numeric keys remain numbers instead of strings.
    """
    return {float(k) if k.replace('.', '', 1).isdigit() else k: v for k, v in data.items()}

def update_excel_mapping(data_dict, user_query):
    prompt = ChatPromptTemplate.from_template("""
    You are analyzing structured data extracted from an Excel sheet.
    The data is provided in dictionary format, where keys represent the content in a cell 
    and values indicate corresponding cell address.
    Additionally, a **user query** is provided. Your task is to:
    1. Identify and count the **distinct tables** in the sheet based on relative address differences.
    2. Find patterns within the data.

    #### Given Data:
    - **Excel Data (Dictionary Format):**  
    {data_dict}

    - **User Query:**  
    {user_query}

    ### **Expected Output Format:**  
    ```plaintext
    Number of tables: [total_count]
    Give a textual description for the input data 
    ```
    """)

    response = model.invoke(prompt.format(
    data_dict=json.dumps(data_dict, indent=4),
    user_query=user_query
    ))

    response_text = response.content
    
    return response

def get_sheet_names(file_path):
    wb = load_workbook(file_path, keep_vba=True)
    return wb.sheetnames



def update_excel_from_json(json_data, excel_file, output_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        return

    sheet = wb[sheet_name]

    for key, cell_refs in json_data.items():
        if isinstance(cell_refs, str):  # If cell_refs is a single string with multiple cell references
            cell_list = [cell.strip() for cell in cell_refs.split(',')]
        elif isinstance(cell_refs, list):  # If stored as a list (alternative format)
            cell_list = cell_refs
        else:
            continue  # Skip unexpected formats

        # Determine whether `key` is a label (string) or value (number)
        value_to_write = key if isinstance(key, (str, bool)) else float(key)

        for cell_ref in cell_list:
            try:
                sheet[cell_ref] = value_to_write
            except ValueError:
                print(f"Invalid cell reference: {cell_ref}")

    wb.save(output_file)
    print(f"Updated Excel saved as: {output_file}")

st.title("Excel Modifier App")
st.write("Upload a `.xlsm` file")

uploaded_file = st.file_uploader("Upload your .xlsm file", type=["xlsm"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name
    sheet_names = get_sheet_names(file_path)
    selected_sheet = st.selectbox("Select a sheet to modify", sheet_names)
    user_command = st.text_input("Enter your modification command')")

    if st.button("Modify Excel"):
        if user_command:
            wb = load_workbook(file_path, data_only = True) 
            
            sheet = wb[selected_sheet]
            json_file = converter(sheet)
            st.write(json_file)
            
            data = {str(k): v for k, v in json_file.items()} ## data is just text
            json_data = json.dumps(data, indent=4) ## this is a json file that contains the sheet info.
            
            text_data = json_to_text(json_file)
            text_file_path = file_path.replace(".xlsm", "_text_for_rag.txt")
            with open(text_file_path, "w") as f:
                f.write(text_data)
            # st.write(f'This is json_data {data}')
            
            with open(text_file_path, "rb") as f:
                st.download_button("📄 Download Text File",f,file_name="excel_mapping_text_for_rag.txt",mime="text/plain")
            with open(text_file_path, "r") as f:
                text = f.read()
            text_splitter = CharacterTextSplitter(chunk_size=20, chunk_overlap=5)
            encoding = tiktoken.encoding_for_model("gpt-4")
            print(f"Original token count: {len(encoding.encode(text))}")

            texts = text_splitter.split_text(text)
            print(f"Token count: {len(encoding.encode(text))}")
            texts = text_splitter.split_text(text)
            docs = [Document(page_content=chunk) for chunk in texts]

            # st.write("Retrieved Chunks:", docs)
            embedding_model = OpenAIEmbeddings()  # or HuggingFaceEmbeddings, etc.
            vectorstore = FAISS.from_documents(docs, embedding_model)
            st.write("Number of documents stored in FAISS vectorstore:", len(vectorstore.index.reconstruct_n(0, vectorstore.index.ntotal)))
            st.write("✅ Number of vectors stored in FAISS index:", vectorstore.index.ntotal)

            qa = RetrievalQA.from_chain_type(llm=model, retriever=vectorstore.as_retriever())

            # query = st.text_input("Ask something about the Excel mapping:")
            # response = update_excel_mapping(data, user_command)
            # print(f'this is what i need {json_updated}')
            # output_file_path = file_path.replace(".xlsm", "_modified.xlsm") 
            # print(output_file_path)
            # result = update_excel_from_json(json_updated, file_path,output_file_path , sheet_name)
            ques = "Explain  the sheet"
            response = qa.run(ques)
            st.write("Response:", response)
            st.write(response)
            st.success(response)
            # query = st.text_input("Ask something about the Excel mapping:")

            # if st.button("Ask"):
            #     if query.strip():
            #         st.write(f"Your query was: {query}")
            #         qa_prompt = f"Answer the following question about the Excel sheet:\n{query}"
            #         response = qa.run(qa_prompt)
            #         st.write("Response:", response)
            # else:
            #     st.warning("⚠️ Please enter a query before clicking Ask.")
           
            # with open(output_file_path, "rb") as f:
            #     st.download_button("Download Modified Excel", f, file_name="modified.xlsm", mime="application/vnd.ms-excel")
            # else:
            #     st.warning("Please enter a command to modify the file.")
