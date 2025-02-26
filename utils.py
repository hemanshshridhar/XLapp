import re
import datetime
import pandas as pd
from dateutil.parser import parse
from openpyxl import load_workbook
import pandas as pd

class IndexColumnConverter:
    def parse_colindex(self, colindex):
        
        result = ""
        while colindex > 0:
            colindex, remainder = divmod(colindex - 1, 26)
            result = chr(65 + remainder) + result
        return result

def get_format(sheet, row, col):
    """Extract formatting from a specific cell"""
    cell = sheet.cell(row=row, column=col)
    format_array = []

    if cell.border.top.style:
        format_array.append('Top Border')
    if cell.border.bottom.style:
        format_array.append('Bottom Border')
    if cell.border.left.style:
        format_array.append('Left Border')
    if cell.border.right.style:
        format_array.append('Right Border')
    if cell.fill.start_color.index not in ['00000000', 'FFFFFFFF']:
        format_array.append('Fill Color')
    if cell.font.bold:
        format_array.append('Font Bold')

    return ', '.join(format_array) if format_array else None

def encode(sheet):
    """Extracts cell values and formatting, and stores them in a DataFrame"""
    converter = IndexColumnConverter()
    markdown = pd.DataFrame(columns=['Address', 'Value', 'Format'])

    for rowindex, row in enumerate(sheet.iter_rows()):
        for colindex, cell in enumerate(row):
            cell_address = converter.parse_colindex(colindex + 1) + str(rowindex + 1)
            cell_value = cell.value
            cell_format = get_format(sheet, rowindex + 1, colindex + 1)

            new_row = pd.DataFrame([[cell_address, cell_value, cell_format]], columns=markdown.columns)
            markdown = pd.concat([markdown, new_row], ignore_index=True)

    return markdown

def guess_datetime_format(string):
    """Try to guess the format of a datetime string"""
    try:
        parsed_date = parse(string, fuzzy=False)
        return parsed_date.strftime('%Y/%m/%d')
    except (ValueError, TypeError):
        return None

def get_category(string):
    """Categorizes the given input string into predefined types."""
    if pd.isna(string):  
        return 'Other'

    if isinstance(string, float):
        return 'Float'

    if isinstance(string, int):
        return 'Integer'

    if isinstance(string, datetime.datetime):
        return 'yyyy/mm/dd'

    string = str(string).strip()  # Convert to string and remove extra spaces

    if re.match(r'^(\+|-)?\d+$', string) or re.match(r'^\d{1,3}(,\d{1,3})*$', string):
        return 'Integer'

    if re.match(r'^[-+]?\d*\.?\d*$', string) or re.match(r'^\d{1,3}(,\d{3})*(\.\d+)?$', string):
        return 'Float'

    if re.match(r'^[-+]?\d*\.?\d*%$', string) or re.match(r'^\d{1,3}(,\d{3})*(\.\d+)?%$', string):
        return 'Percentage'

    if re.match(r'^[-+]?[$]\d*\.?\d{2}$', string) or re.match(r'^[-+]?[$]\d{1,3}(,\d{3})*(\.\d{2})?$', string):
        return 'Currency'

    if re.match(r'\b-?[1-9](?:\.\d+)?[Ee][-+]?\d+\b', string):
        return 'Scientific Notation'

    if re.match(r"^((([!#$%&'*+\-/=?^_`{|}~\w])|([!#$%&'*+\-/=?^_`{|}~\w][!#$%&'*+\-/=?^_`{|}~\.\w]{0,}[!#$%&'*+\-/=?^_`{|}~\w]))[@]\w+([-.]\w+)*\.\w+([-.]\w+)*)$", string):
        return 'Email'

    datetime_format = guess_datetime_format(string)
    if datetime_format:
        return datetime_format

    return 'Other'

def inverted_category(markdown):
  dictionary = {}
  for _, i in markdown.iterrows():
    dictionary[i['Value']] = i['Category']
    return dictionary
def inverted_index(markdown: pd.DataFrame):


    def extract_row_col(address):
        """Extracts column and row from an address like '$A$1' or 'A1'."""
        match = re.match(r"(\$?([A-Z]+))(\$?(\d+))", address)
        if match:
            col, row = match.group(2), int(match.group(4))
            return col, row
        return None, None

    def combine_cells(addresses):
        """Combines addresses into a range when they belong to the same row and have consecutive columns."""
        grouped = {}

        for address in addresses:
            col, row = extract_row_col(address)
            if row is None:  # Skip invalid addresses
                continue
            if row in grouped:
                grouped[row].append(col)
            else:
                grouped[row] = [col]

        def col_to_num(col):
            """Convert column letters (e.g., 'A', 'B', 'AA') to numbers."""
            num = 0
            for c in col:
                num = num * 26 + (ord(c) - ord('A') + 1)
            return num

        def num_to_col(num):
            """Convert numbers back to column letters (e.g., 1 -> 'A', 27 -> 'AA')."""
            col = ""
            while num > 0:
                num -= 1
                col = chr(num % 26 + ord('A')) + col
                num //= 26
            return col

        combined = []
        for row, cols in grouped.items():
            cols = sorted(cols, key=col_to_num)
            start = cols[0]
            prev = start

            ranges = []
            for i in range(1, len(cols)):
                if col_to_num(cols[i]) == col_to_num(prev) + 1:
                    prev = cols[i]
                else:
                    ranges.append(f"{start}{row}:{prev}{row}" if start != prev else f"{start}{row}")
                    start = cols[i]
                    prev = start

        
            ranges.append(f"{start}{row}:{prev}{row}" if start != prev else f"{start}{row}")
            combined.append(", ".join(ranges))

        return ", ".join(combined)

    dictionary = {}
    for _, row in markdown.iterrows():
        if row['Value'] in dictionary:
            dictionary[row['Value']].append(row['Address'])
        else:
            dictionary[row['Value']] = [row['Address']]



    dictionary = {k: v for k, v in dictionary.items() if not pd.isna(k)}
    dictionary = {k: combine_cells(v) for k, v in dictionary.items()}

    return dictionary
def update_excel_from_json(json_data, excel_file, output_file, sheet_name):

    wb = load_workbook(excel_file, keep_vba=True)

   
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        return

    sheet = wb[sheet_name]

   
    for key, cell_refs in json_data.items():

        if isinstance(cell_refs, str):

            cell_list = [cell.strip() for cell in cell_refs.split(',')]

            for cell_ref in cell_list:
                try:
                    sheet[cell_ref] = key 
                except ValueError:
                    print(f"Invalid cell reference: {cell_ref}")

        else:
            
            if isinstance(cell_refs, str):
                cell_list = [cell.strip() for cell in cell_refs.split(',')]
                for cell_ref in cell_list:
                    try:
                        sheet[cell_ref] = key  
                    except ValueError:
                        print(f"Invalid cell reference: {cell_ref}")


    wb.save(output_file)
def converter(sheet):
    df = encode(sheet)
    df['Category'] = df['Value'].apply(lambda x: get_category(x))
    result2  = inverted_index(df)
    return result2
