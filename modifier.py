from langchain_community.llms import Ollama
import streamlit as st
import os
from dotenv import load_dotenv
import json
import re
import tempfile
from openpyxl import load_workbook
from utils import  converter
from fastapi import FastAPI
from langchain.chat_models import ChatOpenAI
import openai
import openpyxl

os.environ["LANGCHAIN_TRACING_V2"] = "true"
api_key = st.secrets["OPENAI_API_KEY"]
langchain_key = st.secrets["LANGCHAIN_API_KEY"]
model = ChatOpenAI(model_name="gpt-4")


def clean_and_parse_json(response):
    """Cleans and parses JSON safely, handling errors."""
    try:
        if not response or not isinstance(response, str):
            raise ValueError("Response is empty or not a string.")

 
        response = re.sub(r"//.*", "", response)


        response = response.replace("'", '"')  


        match = re.search(r"\{.*\}", response, re.DOTALL)
        if match:
            return json.loads(match.group(0))  

    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
    except ValueError as e:
        print(f"Value error: {e}")

    return None 

# def update_excel_mapping(data_dict, user_query):
#     prompt = ChatPromptTemplate.from_template("""
#     You are analyzing structured data extracted from an Excel sheet. The data is provided in dictionary format, where keys represent the content in a cell and values indicate corresponding cell address.
#     Additionally, a **user query** is provided. Your task is to:
#     According to the user query make changes in the data provieded and return the data as it is without any addtion of comments
#     the integer value are the ones not enclosed in " " inverted commas so dont add " " to the integers in the rsponse
    
#     #### Given Data:
#     - **Excel Data (Dictionary Format):**  
#     {data_dict}

#     - **User Query:**  
#     {user_query}

#     ### **Expected Output Format:**  
#     ```plaintext
#     the updated dictionary in the same format.
#     ```
#     """)

#     response = model.invoke(prompt.format(
#     data_dict=json.dumps(data_dict, indent=4),
#     user_query=user_query
#     ))

#     response_text = response.content
#     st.write("Response:", response_text)
#     return response

# def update_excel_mapping(data_dict, user_query):
#     # Define the structured prompt
#     prompt = ChatPromptTemplate.from_template("""
#     You are analyzing structured data extracted from an Excel sheet. The data is provided in dictionary format, where:
#     - Keys represent the content in a cell.
#     - Values represent corresponding cell addresses.
    
#     The user has requested a modification to this data. Your task is to:
#     - Modify the data **precisely as requested** by the user.
#     - Maintain the original format of the dictionary.
#     - **Do not add any explanations, comments, or extra text.**
#     - **Do not enclose integers in quotes.**
    
#     #### Given Data:
#     ```json
#     {data_dict}
#     ```
    
#     #### User Query:
#     {user_query}
    
#     #### Expected Output:
#     Return the updated dictionary **in valid JSON format only**.
#     """)

#     # Call the GPT-4 API
#     response = model.invoke(prompt.format(
#         data_dict=json.dumps(data_dict, separators=(',', ':')),  # Compact JSON
#         user_query=user_query
#     ))

#     # Extract the response and ensure it is valid JSON
#     response_text = response.content.strip()  # Remove any accidental whitespace
#     try:
#         updated_data = json.loads(response_text)  # Convert response back to dictionary
#         st.write("Updated Dictionary:", updated_data)
#         return updated_data
#     except json.JSONDecodeError:
#         st.error("Error: The model did not return valid JSON.")
#         return None


# def update_excel_mapping(data_dict, user_query):
#     # Define the structured prompt
#     prompt = f"""
#     You are analyzing structured data extracted from an Excel sheet. The data is provided in dictionary format, where:
#     - Keys represent the content in a cell.
#     - Values represent corresponding cell addresses.
    
#     The user has requested a modification to this data. Your task is to:
#     - Modify the data **precisely as requested** by the user.
#     - Maintain the original format of the dictionary.
#     - **Do not add any explanations, comments, or extra text.**
#     - **Do not enclose integers in quotes.**
    
#     #### Given Data:
#     ```json
#     {json.dumps(data_dict, separators=(',', ':'))}
#     ```
    
#     #### User Query:
#     {user_query}
    
#     #### Expected Output:
#     Return the updated dictionary **in valid JSON format only**.
#     """

    
#     response = openai.ChatCompletion.create(
#     model="gpt-4",
#     messages=[{"role": "system", "content": "You are an AI assistant."},
#                       {"role": "user", "content": prompt}],
#     temperature=0.2  
#     )
#     response_text = response["choices"][0]["message"]["content"].strip()
#     updated_data = json.loads(response_text)
#     st.write("Updated Dictionary:", updated_data)
#     return updated_data


client = openai.Client(api_key=openai.api_key)

# def update_excel_mapping(data_dict, user_query):
#     prompt = f"""
#     You are analyzing structured data extracted from an Excel sheet. The data is provided in dictionary format, where:
#     - Keys represent the content in a cell.
#     - Values represent corresponding cell addresses.
    
#     The user has requested a modification to this data. Your task is to:
#     - Modify the data **precisely as requested** by the user.
#     - Maintain the original format of the dictionary.
#     - **treat integers as numbers**
#     - **Do not enclose integers in quotes.**
    
#     #### Given Data:
#     ```json
#     {json.dumps(data_dict, separators=(',', ':'))}
#     ```
    
#     #### User Query:
#     {user_query}
    
#     #### Expected Output:
#     Return the updated dictionary **in valid JSON format only**.
#     """

#     response = client.chat.completions.create(
#             model="gpt-4-turbo",
#             messages=[
#                 {"role": "system", "content": "You are an AI assistant."},
#                 {"role": "user", "content": prompt}
#             ],
#             temperature=0.2
#         )

#         # Extract response text
#     response_text = response.choices[0].message.content.strip()

#         # Convert response to dictionary
    

#         # Display the result
#     st.write("Updated Dictionary:", response_text)
#     return response_text

def parse_json_response(response_text):
    if not response_text.strip():
        raise ValueError("Error: GPT response is empty. Cannot parse JSON.")

    # Remove triple backticks (` ```json ` or ` ``` `)
    cleaned_response = re.sub(r"```(?:json)?\n?|```", "", response_text).strip()

    try:
        return json.loads(cleaned_response, parse_float=float, parse_int=int)
    except json.JSONDecodeError as e:
        raise ValueError(f"Error parsing JSON: {e}. Cleaned Response was: {repr(cleaned_response)}")



def enforce_numeric_keys(data):
    """
    Ensures that numeric keys remain numbers instead of strings.
    """
    return {float(k) if k.replace('.', '', 1).isdigit() else k: v for k, v in data.items()}

def update_excel_mapping(data_dict, user_query):
    prompt = f"""
    You are analyzing structured data extracted from an Excel sheet.
    The data is provided in dictionary format, where:
    - Keys represent the content in a cell.
    - Values represent corresponding cell addresses.
    
    The user has requested a modification to this data. Your task is to:
    - Modify the data **precisely as requested** by the user.
    - Maintain the original format of the dictionary.
    - **Treat integers as numbers**
    - **Do not enclose integers in quotes.**
    -**Do not use triple backticks (` ``` `) or markdown.**
    
    #### Given Data:
    ```json
    {json.dumps(data_dict, separators=(',', ':'))}
    ```
    
    #### User Query:
    {user_query}
    
    #### Expected Output:
    Return the updated dictionary **in valid JSON format only**.
    """

    response = client.chat.completions.create(
        model="gpt-4-turbo",
        messages=[
            {"role": "system", "content": "You are an AI assistant."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.2
    )

    # Extract response text
    response_text = response.choices[0].message.content.strip()

    # Convert response to dictionary with proper numeric formatting
    parsed_dict = parse_json_response(response_text)
    cleaned_dict = enforce_numeric_keys(parsed_dict)

    st.write("Updated Dictionary:", cleaned_dict)
    st.json(json.dumps(cleaned_dict, indent=4))
    return cleaned_dict





def update_excel_from_json(json_data, excel_file, output_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        return

    sheet = wb[sheet_name]

    for key, cell_refs in json_data.items():
        if isinstance(cell_refs, str):  # If cell_refs is a single string with multiple cell references
            cell_list = [cell.strip() for cell in cell_refs.split(',')]
        elif isinstance(cell_refs, list):  # If stored as a list (alternative format)
            cell_list = cell_refs
        else:
            continue  # Skip unexpected formats

        # Determine whether `key` is a label (string) or value (number)
        value_to_write = key if isinstance(key, (str, bool)) else float(key)

        for cell_ref in cell_list:
            try:
                sheet[cell_ref] = value_to_write
            except ValueError:
                print(f"Invalid cell reference: {cell_ref}")

    wb.save(output_file)
    print(f"Updated Excel saved as: {output_file}")

st.title("Excel Modifier App")
st.write("Upload a `.xlsm` file and modify it using natural language.")

uploaded_file = st.file_uploader("Upload your .xlsm file", type=["xlsm"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name

    user_command = st.text_input("Enter your modification command')")

    if st.button("Modify Excel"):
        if user_command:
            wb = load_workbook(file_path, keep_vba=True) 
            sheet_name = "Model Inputs"
            sheet = wb[sheet_name]
            json_file = converter(sheet)
            
            
            data_dict_fixed = {str(k): v for k, v in json_file.items()}
            json_updated = update_excel_mapping(data_dict_fixed, user_command)
            print(f'this is what i need {json_updated}')
            output_file_path = file_path.replace(".xlsm", "_modified.xlsm") 
            print(output_file_path)
            result = update_excel_from_json(json_updated, file_path,output_file_path , sheet_name)
            st.success(result)

           
            with open(output_file_path, "rb") as f:
                st.download_button("Download Modified Excel", f, file_name="modified.xlsm", mime="application/vnd.ms-excel")
        else:
            st.warning("Please enter a command to modify the file.")
