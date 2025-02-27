import streamlit as st
import os
import json
import re
import tempfile
import torch
from openpyxl import load_workbook
from transformers import AutoModelForCausalLM, AutoTokenizer
from dotenv import load_dotenv
from utils import converter  
from huggingface_hub import login
load_dotenv()

os.environ["HUGGINGFACE_TOKEN"] = "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"
login(token=os.getenv("HUGGINGFACE_TOKEN"))

MODEL_NAME = "mistralai/Mistral-7B-Instruct-v0.1"

@st.cache_resource()
def load_hf_model():
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    model = AutoModelForCausalLM.from_pretrained(MODEL_NAME, torch_dtype=torch.float16, device_map="auto")
    return tokenizer, model

tokenizer, model = load_hf_model()

def generate_response(prompt):
    inputs = tokenizer(prompt, return_tensors="pt").to("cuda")
    output = model.generate(**inputs, max_length=500)
    response = tokenizer.decode(output[0], skip_special_tokens=True)
    return response

def clean_and_parse_json(response):
    """Cleans and parses JSON safely, handling errors."""
    try:
        if not response or not isinstance(response, str):
            raise ValueError("Response is empty or not a string.")

        response = re.sub(r"//.*", "", response)
        response = response.replace("'", '"')  

        match = re.search(r"\{.*\}", response, re.DOTALL)
        if match:
            return json.loads(match.group(0)) 

    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
    except ValueError as e:
        print(f"Value error: {e}")

    return None 

def update_excel_mapping(data_dict, user_query):
    prompt = f"""
    You are analyzing structured data extracted from an Excel sheet. The data is provided in dictionary format, where keys represent category names and values indicate corresponding cell locations.

    Additionally, a **user query** is provided. Your task is to:
    1. Identify and count the **distinct tables** in the sheet based on realtive address difference whereever a large differencce in Row number is there, a new table starts .
    2. Ensure that **each table is defined by a set of related rows**.
    3. Find patterns within the data.

    #### Given Data:
    - **Excel Data (Dictionary Format):**  
    {json.dumps(data_dict, indent=4)}
    - **User Query:**  
    "{user_query}"

    ### **Expected Output Format:**  
    ```plaintext
    Number of tables: [total_count]
    Table Names: [List of identified table names]

    Provide a concise and accurate response based on the given data.
    """
    response = generate_response(prompt) 
    print(response)
    st.write("Raw Response from Hugging Face Model:", response)

    return response

def update_excel_from_json(json_data, excel_file, output_file, sheet_name):
    wb = load_workbook(excel_file, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
        return

    sheet = wb[sheet_name]

    for key, cell_refs in json_data.items():
        if isinstance(cell_refs, str):
            cell_list = [cell.strip() for cell in cell_refs.split(',')]
            for cell_ref in cell_list:
                try:
                    sheet[cell_ref] = key  
                except ValueError:
                    print(f"Invalid cell reference: {cell_ref}")

    wb.save(output_file)

# Streamlit UI
st.title("Excel Sheet Analyzer using Hugging Face 🚀")
st.write("Upload a `.xlsm` file and modify it using natural language.")

uploaded_file = st.file_uploader("Upload your .xlsm file", type=["xlsm"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
        temp_file.write(uploaded_file.read())
        file_path = temp_file.name

    user_command = st.text_input("Enter your query")

    if st.button("Get answer"):
        st.write("Excel sheet analysis")
        if user_command:
            wb = load_workbook(file_path, keep_vba=True) 
            sheet = wb['Model Inputs']
            json_file = converter(sheet)
            sheet_name = "Model Inputs"

            data_dict_fixed = {str(k): v for k, v in json_file.items()}
            answer = update_excel_mapping(json_file, user_command)

            st.write("Excel sheet analysis")
        else:
            st.warning("Please give a query")
